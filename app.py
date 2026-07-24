import os, csv, io, json, uuid, hashlib, base64, socket, re, unicodedata, sqlite3, calendar
from datetime import datetime, date
from functools import wraps
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, abort, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, case, or_, inspect, text
from werkzeug.utils import secure_filename

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
except ImportError:
    service_account = build = MediaFileUpload = MediaIoBaseDownload = None
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

APP_VERSION='15.2'
BASE=Path(__file__).resolve().parent
UPLOAD=BASE/'static'/'uploads'; UPLOAD.mkdir(parents=True,exist_ok=True)
DRIVE_CONFIG_FILE=BASE/'config_drive.json'

IMPORT_DIR=BASE/'data'/'imports'; IMPORT_DIR.mkdir(parents=True,exist_ok=True)

def norm_header(value):
    text=unicodedata.normalize('NFKD',str(value or '')).encode('ascii','ignore').decode().lower().strip()
    return re.sub(r'[^a-z0-9]+',' ',text).strip()

COLUMN_ALIASES={
 'sat':['sat','n sat','numero sat','numero do chamado','chamado','protocolo','id chamado','codigo'],
 'empreendimento':['empreendimento','condominio','obra','projeto','residencial'],
 'data_recebido':['data recebido','data recebimento','data abertura','abertura','data do chamado','data'],
 'solicitante':['solicitante','morador','cliente','nome morador','proprietario','nome cliente'],
 'unidade':['unidade','apartamento','apto','apt','torre apartamento','bloco unidade'],
 'problema':['problema','descricao','solicitacao','assunto','defeito','ocorrencia','relato'],
 'contato':['contato','telefone','celular','whatsapp','fone'],
 'status':['status','situacao','andamento'],
 'classificacao':['classificacao','procedencia','resultado'],
 'categoria':['categoria','tipo','sistema','item'],
 'data_entrega':['data entrega','entrega empreendimento','data de entrega'],
 'responsavel':['responsavel','tecnico','equipe','colaborador'],
 'observacoes':['observacoes','observacao','obs','comentarios']
}
ALIAS_LOOKUP={norm_header(a):field for field,aliases in COLUMN_ALIASES.items() for a in aliases}

def excel_date(value):
    if value is None: return ''
    if isinstance(value,(datetime,date)): return value.strftime('%Y-%m-%d')
    text=str(value).strip()
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%d/%m/%y'):
        try: return datetime.strptime(text,fmt).strftime('%Y-%m-%d')
        except ValueError: pass
    return text

def read_import_file(file_storage):
    filename=(file_storage.filename or '').lower()
    rows=[]
    if filename.endswith('.csv'):
        raw=file_storage.read()
        text=raw.decode('utf-8-sig',errors='replace')
        sample=text[:4096]
        try: dialect=csv.Sniffer().sniff(sample,delimiters=';,\t,')
        except Exception: dialect=csv.excel; dialect.delimiter=';'
        rows=list(csv.reader(io.StringIO(text),dialect))
    elif filename.endswith(('.xlsx','.xlsm')):
        wb=load_workbook(file_storage,read_only=True,data_only=True)
        ws=wb.active
        rows=[list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise ValueError('Envie uma planilha .xlsx, .xlsm ou .csv.')
    if not rows: raise ValueError('A planilha está vazia.')
    header_idx=0; best=-1
    for i,row in enumerate(rows[:12]):
        score=sum(1 for v in row if norm_header(v) in ALIAS_LOOKUP)
        if score>best: best=score; header_idx=i
    headers=[norm_header(v) for v in rows[header_idx]]
    mapping={i:ALIAS_LOOKUP[h] for i,h in enumerate(headers) if h in ALIAS_LOOKUP}
    if 'problema' not in mapping.values() and 'sat' not in mapping.values():
        raise ValueError('Não foi possível reconhecer as colunas. Use o modelo disponível no sistema.')
    parsed=[]
    for n,row in enumerate(rows[header_idx+1:],start=header_idx+2):
        if not any(v not in (None,'') for v in row): continue
        item={'_linha':n}
        for i,field in mapping.items():
            value=row[i] if i<len(row) else ''
            item[field]=excel_date(value) if field in ('data_recebido','data_entrega') else str(value or '').strip()
        item.setdefault('data_recebido',date.today().isoformat())
        item.setdefault('status','Aberto')
        item.setdefault('categoria','')
        item['_erro']='' if (item.get('sat') or item.get('problema')) else 'Informe SAT ou descrição do problema.'
        parsed.append(item)
    return parsed, mapping
DB_URL=os.environ.get('DATABASE_URL') or f"sqlite:///{BASE/'data'/'gestao10.db'}"
if DB_URL.startswith('postgres://'): DB_URL='postgresql://'+DB_URL[len('postgres://'):]

app=Flask(__name__)
app.secret_key=os.environ.get('SECRET_KEY','troque-esta-chave-no-render-brsmartsat12')
app.config.update(SQLALCHEMY_DATABASE_URI=DB_URL,SQLALCHEMY_TRACK_MODIFICATIONS=False,MAX_CONTENT_LENGTH=40*1024*1024)
db=SQLAlchemy(app)

def ph(s): return hashlib.sha256(s.encode()).hexdigest()
def check(h,s): return h==ph(s)

class Usuario(db.Model):
    __tablename__='usuarios'; id=db.Column(db.Integer,primary_key=True); nome=db.Column(db.String(120)); email=db.Column(db.String(180),unique=True,nullable=False); senha=db.Column(db.String(128)); perfil=db.Column(db.String(30),default='Equipe'); construtora=db.Column(db.String(150)); ativo=db.Column(db.Boolean,default=True)
class Construtora(db.Model):
    __tablename__='construtoras'; id=db.Column(db.Integer,primary_key=True); nome=db.Column(db.String(150),unique=True,nullable=False); ativo=db.Column(db.Boolean,default=True)
class Empreendimento(db.Model):
    __tablename__='empreendimentos'; id=db.Column(db.Integer,primary_key=True); nome=db.Column(db.String(150),unique=True); construtora=db.Column(db.String(150),default='Prestes'); data_entrega=db.Column(db.String(10)); observacoes=db.Column(db.Text)
class Chamado(db.Model):
    __tablename__='chamados'; id=db.Column(db.Integer,primary_key=True); sat=db.Column(db.String(80)); origem=db.Column(db.String(30),default='Prestes'); construtora=db.Column(db.String(150),default='Prestes'); empreendimento=db.Column(db.String(150)); data_recebido=db.Column(db.String(10)); solicitante=db.Column(db.String(180)); unidade=db.Column(db.String(100)); problema=db.Column(db.Text); contato=db.Column(db.String(120)); status=db.Column(db.String(80),default='Aberto'); classificacao=db.Column(db.String(100)); categoria=db.Column(db.String(220)); data_entrega=db.Column(db.String(10)); analise_garantia=db.Column(db.String(120)); fundamentacao=db.Column(db.Text); responsavel=db.Column(db.String(180)); observacoes=db.Column(db.Text); atualizado_em=db.Column(db.DateTime,default=datetime.utcnow,onupdate=datetime.utcnow)
class Garantia(db.Model):
    __tablename__='garantias'; id=db.Column(db.Integer,primary_key=True); empreendimento=db.Column(db.String(150)); item=db.Column(db.String(220)); prazo_meses=db.Column(db.Integer); classificacao_padrao=db.Column(db.String(100)); descricao=db.Column(db.Text); fonte=db.Column(db.Text)
class Foto(db.Model):
    __tablename__='fotos'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); arquivo=db.Column(db.String(255)); descricao=db.Column(db.String(255)); criado_em=db.Column(db.DateTime,default=datetime.utcnow); latitude=db.Column(db.Float); longitude=db.Column(db.Float); precisao=db.Column(db.Float); capturado_em=db.Column(db.DateTime); usuario=db.Column(db.String(180)); dispositivo=db.Column(db.String(255)); drive_file_id=db.Column(db.String(180)); drive_web_link=db.Column(db.Text); drive_folder_link=db.Column(db.Text); armazenamento=db.Column(db.String(30),default='Local')
class Assinatura(db.Model):
    __tablename__='assinaturas'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); arquivo=db.Column(db.String(255)); nome_cliente=db.Column(db.String(180)); documento=db.Column(db.String(80)); observacao=db.Column(db.Text); criado_em=db.Column(db.DateTime,default=datetime.utcnow); latitude=db.Column(db.Float); longitude=db.Column(db.Float); precisao=db.Column(db.Float); assinado_em=db.Column(db.DateTime); usuario=db.Column(db.String(180)); dispositivo=db.Column(db.String(255))
class Material(db.Model):
    __tablename__='materiais'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); item=db.Column(db.String(180)); quantidade=db.Column(db.String(80)); status=db.Column(db.String(80)); observacoes=db.Column(db.Text)
class Financeiro(db.Model):
    __tablename__='financeiro'; id=db.Column(db.Integer,primary_key=True); data=db.Column(db.String(10)); tipo=db.Column(db.String(20)); descricao=db.Column(db.Text); origem=db.Column(db.String(180)); construtora=db.Column(db.String(150)); empreendimento=db.Column(db.String(150)); categoria=db.Column(db.String(120)); valor=db.Column(db.Float,default=0); status=db.Column(db.String(50)); forma_pagamento=db.Column(db.String(50)); vencimento=db.Column(db.String(10)); documento=db.Column(db.String(80)); observacoes=db.Column(db.Text)
class Agenda(db.Model):
    __tablename__='agenda'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); construtora=db.Column(db.String(150)); empreendimento=db.Column(db.String(150)); lider=db.Column(db.String(180)); lider_id=db.Column(db.Integer); lider_email=db.Column(db.String(180)); equipe=db.Column(db.String(180)); equipe_id=db.Column(db.Integer); equipe_email=db.Column(db.String(180)); prazo_contato=db.Column(db.String(10)); data_agendada=db.Column(db.String(10)); hora=db.Column(db.String(8)); periodo=db.Column(db.String(20)); status=db.Column(db.String(60),default='Aguardando contato'); tipo_fluxo=db.Column(db.String(40)); contato_status=db.Column(db.String(60)); observacao_interna=db.Column(db.Text); observacao_construtora=db.Column(db.Text); criado_por=db.Column(db.String(180)); agendado_por=db.Column(db.String(180)); lider_visualizou=db.Column(db.Boolean,default=False); lider_confirmou=db.Column(db.Boolean,default=False); confirmado_em=db.Column(db.DateTime); criado_em=db.Column(db.DateTime,default=datetime.utcnow); atualizado_em=db.Column(db.DateTime,default=datetime.utcnow,onupdate=datetime.utcnow)
class Historico(db.Model):
    __tablename__='historico'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); usuario=db.Column(db.String(180)); acao=db.Column(db.String(220)); criado_em=db.Column(db.DateTime,default=datetime.utcnow)
class Auditoria(db.Model):
    __tablename__='auditoria'; id=db.Column(db.Integer,primary_key=True); modulo=db.Column(db.String(50)); registro_id=db.Column(db.Integer); usuario=db.Column(db.String(180)); acao=db.Column(db.String(80)); detalhes=db.Column(db.Text); criado_em=db.Column(db.DateTime,default=datetime.utcnow)

class Atendimento(db.Model):
    __tablename__='atendimentos'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id')); tipo=db.Column(db.String(30)); criado_em=db.Column(db.DateTime,default=datetime.utcnow); latitude=db.Column(db.Float); longitude=db.Column(db.Float); precisao=db.Column(db.Float); usuario=db.Column(db.String(180)); dispositivo=db.Column(db.String(255)); observacao=db.Column(db.Text)
class RelatorioTecnico(db.Model):
    __tablename__='relatorios_tecnicos'; id=db.Column(db.Integer,primary_key=True); chamado_id=db.Column(db.Integer,db.ForeignKey('chamados.id'),unique=True); diagnostico=db.Column(db.Text); servico_executado=db.Column(db.Text); testes_realizados=db.Column(db.Text); conclusao=db.Column(db.Text); atualizado_em=db.Column(db.DateTime,default=datetime.utcnow,onupdate=datetime.utcnow); usuario=db.Column(db.String(180))

def ensure_columns():
    # Migração simples para instalações que já possuem o banco das versões anteriores.
    required={
      'fotos': {'latitude':'FLOAT','longitude':'FLOAT','precisao':'FLOAT','capturado_em':'TIMESTAMP','usuario':'VARCHAR(180)','dispositivo':'VARCHAR(255)','drive_file_id':'VARCHAR(180)','drive_web_link':'TEXT','drive_folder_link':'TEXT','armazenamento':"VARCHAR(30) DEFAULT 'Local'"},
      'assinaturas': {'latitude':'FLOAT','longitude':'FLOAT','precisao':'FLOAT','assinado_em':'TIMESTAMP','usuario':'VARCHAR(180)','dispositivo':'VARCHAR(255)'},
      'empreendimentos': {'construtora':"VARCHAR(150) DEFAULT 'Prestes'"},
      'chamados': {'origem':"VARCHAR(30) DEFAULT 'Prestes'",'construtora':"VARCHAR(150) DEFAULT 'Prestes'"},
      'financeiro': {'origem':'VARCHAR(180)','construtora':'VARCHAR(150)','vencimento':'VARCHAR(10)','documento':'VARCHAR(80)','observacoes':'TEXT'},
      'usuarios': {'construtora':'VARCHAR(150)'},
      'agenda': {'lider_visualizou':'BOOLEAN DEFAULT 0','lider_confirmou':'BOOLEAN DEFAULT 0','lider_id':'INTEGER','lider_email':'VARCHAR(180)','equipe_id':'INTEGER','equipe_email':'VARCHAR(180)','tipo_fluxo':'VARCHAR(40)','agendado_por':'VARCHAR(180)','confirmado_em':'TIMESTAMP'}
    }
    inspector=inspect(db.engine)
    tables=set(inspector.get_table_names())
    for table,cols in required.items():
        if table not in tables: continue
        existing={c['name'] for c in inspector.get_columns(table)}
        for col,typ in cols.items():
            if col not in existing:
                db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} {typ}'))
        db.session.commit()

def local_drive_root():
    raw=os.environ.get('GOOGLE_DRIVE_LOCAL_ROOT','').strip()
    if not raw and DRIVE_CONFIG_FILE.exists():
        try:
            raw=str(json.loads(DRIVE_CONFIG_FILE.read_text(encoding='utf-8')).get('local_root') or '').strip()
        except Exception:
            raw=''
    if not raw:
        return None
    path=Path(raw).expanduser()
    return path if path.exists() and path.is_dir() else None

def api_drive_configured():
    return bool(os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON') and os.environ.get('GOOGLE_DRIVE_ROOT_FOLDER_ID') and build and service_account)

def drive_configured():
    return bool(local_drive_root() or api_drive_configured())

def drive_mode():
    if local_drive_root(): return 'Google Drive para computador'
    if api_drive_configured(): return 'Google Drive API'
    return 'Não configurado'

def drive_service():
    if not api_drive_configured():
        return None
    raw=os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON','').strip()
    try:
        info=json.loads(raw)
    except json.JSONDecodeError:
        # Também aceita caminho local para facilitar testes no computador.
        path=Path(raw)
        if not path.exists():
            raise RuntimeError('GOOGLE_SERVICE_ACCOUNT_JSON inválido: informe o JSON completo ou o caminho do arquivo.')
        info=json.loads(path.read_text(encoding='utf-8'))
    creds=service_account.Credentials.from_service_account_info(info,scopes=['https://www.googleapis.com/auth/drive'])
    return build('drive','v3',credentials=creds,cache_discovery=False)

def drive_safe_name(value, fallback='Sem nome'):
    text=re.sub(r'[\\/:*?"<>|]+','-',str(value or '').strip())
    return text[:120] or fallback

def drive_find_or_create_folder(service, name, parent_id):
    safe=drive_safe_name(name)
    escaped=safe.replace("'", "\\'")
    query=f"name = '{escaped}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false and '{parent_id}' in parents"
    result=service.files().list(q=query,spaces='drive',fields='files(id,name,webViewLink)',pageSize=10,includeItemsFromAllDrives=True,supportsAllDrives=True).execute()
    files=result.get('files',[])
    if files:
        folder=files[0]
    else:
        folder=service.files().create(body={'name':safe,'mimeType':'application/vnd.google-apps.folder','parents':[parent_id]},fields='id,name,webViewLink',supportsAllDrives=True).execute()
    folder.setdefault('webViewLink',f"https://drive.google.com/drive/folders/{folder['id']}")
    return folder

def _sat_number(chamado):
    return re.sub(r'\D+','',str(chamado.sat or chamado.id))

def local_drive_locate_sat_folder(chamado):
    root=local_drive_root()
    if not root: return None
    sat_num=_sat_number(chamado)
    if not sat_num: return None
    # Procura em qualquer nível por pasta cujo nome contenha exatamente o número da SAT.
    # Exemplos aceitos: "29014", "SAT 29014", "29014 - VISTORIA".
    matches=[]
    for current, dirs, files in os.walk(root):
        # evita pastas técnicas do Google Drive e acelera a busca
        dirs[:]=[d for d in dirs if not d.startswith('.') and d.lower() not in {'lost+found'}]
        for d in dirs:
            nums=re.findall(r'\d+',d)
            if sat_num in nums:
                path=Path(current)/d
                score=0
                emp=unicodedata.normalize('NFKD',str(chamado.empreendimento or '')).encode('ascii','ignore').decode().lower()
                path_norm=unicodedata.normalize('NFKD',str(path)).encode('ascii','ignore').decode().lower()
                if emp and emp in path_norm: score+=10
                if d.strip()==sat_num: score+=5
                if d.lower().startswith('sat '): score+=4
                matches.append((score,len(str(path)),path))
    if not matches: return None
    matches.sort(key=lambda x:(-x[0],x[1]))
    return matches[0][2]

def local_drive_image_files(folder):
    exts={'.jpg','.jpeg','.png','.webp','.gif','.bmp','.heic','.heif'}
    for current, dirs, files in os.walk(folder):
        dirs[:]=[d for d in dirs if not d.startswith('.')]
        for name in files:
            path=Path(current)/name
            if path.suffix.lower() in exts:
                yield path

def local_stage_name(file_path, sat_folder):
    try:
        rel=file_path.relative_to(sat_folder)
        if len(rel.parts)>1:
            raw=rel.parts[0]
        else:
            raw=sat_folder.name
    except Exception:
        raw='Vistoria'
    low=raw.lower()
    if 'antes' in low: return 'Antes do serviço'
    if 'durante' in low: return 'Durante do serviço' if False else 'Durante o serviço'
    if 'depois' in low or 'final' in low: return 'Depois do serviço'
    if 'material' in low: return 'Material utilizado'
    if 'vistoria' in low: return 'Vistoria'
    return 'Vistoria'

def sync_local_drive_photos(chamado):
    sat_folder=local_drive_locate_sat_folder(chamado)
    if not sat_folder:
        return {'importadas':0,'existentes':0,'pasta':None,'modo':'local'}
    imported=existing=0
    for item in local_drive_image_files(sat_folder):
        token='local:'+hashlib.sha256(str(item.resolve()).encode('utf-8')).hexdigest()
        if Foto.query.filter_by(chamado_id=chamado.id,drive_file_id=token).first():
            existing+=1; continue
        try: created=datetime.fromtimestamp(item.stat().st_mtime)
        except Exception: created=datetime.utcnow()
        db.session.add(Foto(chamado_id=chamado.id,arquivo='',descricao=local_stage_name(item,sat_folder),criado_em=created,capturado_em=created,usuario='Google Drive',drive_file_id=token,drive_web_link='',drive_folder_link=str(sat_folder),armazenamento='Google Drive para computador'))
        imported+=1
    if imported: db.session.commit()
    return {'importadas':imported,'existentes':existing,'pasta':{'id':str(sat_folder),'name':sat_folder.name,'webViewLink':''},'modo':'local'}

def upload_photo_to_drive(local_path, chamado, etapa, original_name):
    root=local_drive_root()
    if root:
        emp=root/drive_safe_name(chamado.empreendimento or 'Sem empreendimento')
        sat=emp/f"SAT {chamado.sat or chamado.id}"
        etapa_folder=sat/drive_safe_name(etapa or 'Fotos')
        etapa_folder.mkdir(parents=True,exist_ok=True)
        timestamp=datetime.now().strftime('%Y%m%d_%H%M%S')
        target=etapa_folder/f"{timestamp}_{drive_safe_name(original_name,'foto.jpg')}"
        import shutil
        shutil.copy2(local_path,target)
        token='local:'+hashlib.sha256(str(target.resolve()).encode('utf-8')).hexdigest()
        return {'file_id':token,'web_link':'','folder_link':str(sat)}
    service=drive_service()
    if not service:
        return None
    root=os.environ['GOOGLE_DRIVE_ROOT_FOLDER_ID'].strip()
    emp=drive_find_or_create_folder(service,chamado.empreendimento or 'Sem empreendimento',root)
    sat_name=f"SAT {chamado.sat or chamado.id}"
    sat=drive_find_or_create_folder(service,sat_name,emp['id'])
    etapa_folder=drive_find_or_create_folder(service,etapa or 'Fotos',sat['id'])
    timestamp=datetime.now().strftime('%Y%m%d_%H%M%S')
    filename=f"{timestamp}_{drive_safe_name(original_name,'foto.jpg')}"
    media=MediaFileUpload(str(local_path),resumable=True)
    uploaded=service.files().create(body={'name':filename,'parents':[etapa_folder['id']]},media_body=media,fields='id,name,webViewLink',supportsAllDrives=True).execute()
    uploaded.setdefault('webViewLink',f"https://drive.google.com/file/d/{uploaded['id']}/view")
    return {'file_id':uploaded['id'],'web_link':uploaded['webViewLink'],'folder_link':sat['webViewLink']}



def drive_find_folder(service, name, parent_id):
    safe=drive_safe_name(name)
    escaped=safe.replace("'", "\\'")
    query=f"name = '{escaped}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false and '{parent_id}' in parents"
    result=service.files().list(q=query,spaces='drive',fields='files(id,name,webViewLink)',pageSize=20,includeItemsFromAllDrives=True,supportsAllDrives=True).execute()
    files=result.get('files',[])
    if not files:
        return None
    folder=files[0]
    folder.setdefault('webViewLink',f"https://drive.google.com/drive/folders/{folder['id']}")
    return folder

def drive_list_children(service, parent_id):
    items=[]
    token=None
    while True:
        result=service.files().list(
            q=f"trashed = false and '{parent_id}' in parents",
            spaces='drive',
            fields='nextPageToken,files(id,name,mimeType,webViewLink,createdTime,modifiedTime,parents)',
            pageSize=1000,
            pageToken=token,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True
        ).execute()
        items.extend(result.get('files',[]))
        token=result.get('nextPageToken')
        if not token:
            break
    return items

def drive_locate_sat_folder(service, chamado):
    root=os.environ['GOOGLE_DRIVE_ROOT_FOLDER_ID'].strip()
    emp=drive_find_folder(service,chamado.empreendimento or 'Sem empreendimento',root)
    if not emp:
        return None
    sat_value=str(chamado.sat or chamado.id).strip()
    candidates=[f'SAT {sat_value}',sat_value,f'Sat {sat_value}']
    for name in candidates:
        folder=drive_find_folder(service,name,emp['id'])
        if folder:
            return folder
    normalized=re.sub(r'\D+','',sat_value)
    for item in drive_list_children(service,emp['id']):
        if item.get('mimeType')!='application/vnd.google-apps.folder':
            continue
        item_num=re.sub(r'\D+','',item.get('name',''))
        if normalized and item_num==normalized:
            item.setdefault('webViewLink',f"https://drive.google.com/drive/folders/{item['id']}")
            return item
    return None

def sync_drive_photos(chamado):
    if local_drive_root():
        return sync_local_drive_photos(chamado)
    service=drive_service()
    if not service:
        return {'importadas':0,'existentes':0,'pasta':None}
    sat_folder=drive_locate_sat_folder(service,chamado)
    if not sat_folder:
        return {'importadas':0,'existentes':0,'pasta':None}
    imported=0
    existing=0
    folders=[('Fotos',sat_folder['id'])]
    for child in drive_list_children(service,sat_folder['id']):
        if child.get('mimeType')=='application/vnd.google-apps.folder':
            folders.append((child.get('name') or 'Fotos',child['id']))
    for etapa,folder_id in folders:
        for item in drive_list_children(service,folder_id):
            mime=item.get('mimeType','')
            if not mime.startswith('image/'):
                continue
            if Foto.query.filter_by(chamado_id=chamado.id,drive_file_id=item['id']).first():
                existing+=1
                continue
            web=item.get('webViewLink') or f"https://drive.google.com/file/d/{item['id']}/view"
            created=None
            try:
                created=datetime.fromisoformat((item.get('createdTime') or '').replace('Z','+00:00')).replace(tzinfo=None)
            except Exception:
                created=datetime.utcnow()
            reg=Foto(
                chamado_id=chamado.id,
                arquivo='',
                descricao=etapa,
                criado_em=created,
                capturado_em=created,
                usuario='Google Drive',
                drive_file_id=item['id'],
                drive_web_link=web,
                drive_folder_link=sat_folder.get('webViewLink'),
                armazenamento='Google Drive'
            )
            db.session.add(reg)
            imported+=1
    if imported:
        db.session.commit()
    return {'importadas':imported,'existentes':existing,'pasta':sat_folder}

def delete_drive_file(file_id):
    if not file_id or not drive_configured():
        return
    if str(file_id).startswith('local:'):
        # Por segurança, arquivos sincronizados de uma pasta local do Drive não são apagados automaticamente.
        return
    try:
        drive_service().files().delete(fileId=file_id,supportsAllDrives=True).execute()
    except Exception:
        app.logger.exception('Não foi possível excluir a foto do Google Drive.')

def parse_location(form):
    try:
        lat=float(form.get('latitude','')); lon=float(form.get('longitude',''))
        acc=float(form.get('precisao') or 0)
        if not (-90 <= lat <= 90 and -180 <= lon <= 180): raise ValueError
        return lat,lon,acc
    except (TypeError,ValueError):
        return None

def parse_client_time(value):
    try:
        return datetime.fromisoformat((value or '').replace('Z','+00:00')).replace(tzinfo=None)
    except Exception:
        return datetime.utcnow()

def migrate_legacy_sqlite_to_database():
    """Copia automaticamente os dados do SQLite legado para o PostgreSQL vazio.

    A migração é idempotente: só roda quando DATABASE_URL aponta para PostgreSQL
    e a tabela de usuários ainda está vazia. O arquivo SQLite precisa existir em
    data/gestao10.db, que já acompanha esta atualização.
    """
    if not DB_URL.startswith('postgresql://'):
        return False
    legacy_path=BASE/'data'/'gestao10.db'
    if not legacy_path.exists() or Usuario.query.first():
        return False

    model_order=[
        Usuario, Construtora, Empreendimento, Chamado, Garantia, Foto,
        Assinatura, Material, Financeiro, Agenda, Historico, Auditoria,
        Atendimento, RelatorioTecnico
    ]
    conn=sqlite3.connect(str(legacy_path))
    conn.row_factory=sqlite3.Row
    migrated=0
    try:
        existing_tables={r['name'] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        for model in model_order:
            table=model.__tablename__
            if table not in existing_tables:
                continue
            model_columns={c.name for c in model.__table__.columns}
            rows=conn.execute(f'SELECT * FROM "{table}"').fetchall()
            for row in rows:
                payload={k:row[k] for k in row.keys() if k in model_columns}
                if not payload:
                    continue
                record_id=payload.get('id')
                if record_id is not None and db.session.get(model, record_id):
                    continue
                db.session.add(model(**payload))
                migrated+=1
            db.session.flush()
        db.session.commit()

        # Corrige as sequências do PostgreSQL após inserir IDs vindos do SQLite.
        for model in model_order:
            table=model.__tablename__
            if 'id' not in {c.name for c in model.__table__.columns}:
                continue
            db.session.execute(text(
                "SELECT setval(pg_get_serial_sequence(:table, 'id'), "
                "COALESCE((SELECT MAX(id) FROM \"" + table + "\"), 1), true)"
            ), {'table': table})
        db.session.commit()
        print(f'[MIGRAÇÃO] {migrated} registros copiados do SQLite para o PostgreSQL.')
        return migrated > 0
    except Exception:
        db.session.rollback()
        raise
    finally:
        conn.close()

def init_db():
    db.create_all()
    ensure_columns()
    migrate_legacy_sqlite_to_database()
    if not Construtora.query.first():
        db.session.add(Construtora(nome='Prestes')); db.session.commit()
    if not Usuario.query.first():
        db.session.add(Usuario(nome='Administrador BR',email='admin@brsolucoes.com.br',senha=ph('123456'),perfil='Administrador'))
        db.session.commit()
    # Contas locais de teste para validar o fluxo completo da agenda.
    if not Usuario.query.filter(Usuario.perfil=='Líder').first():
        db.session.add(Usuario(nome='Líder BR',email='lider@brsolucoes.com.br',senha=ph('123456'),perfil='Líder',ativo=True))
    if not Usuario.query.filter(Usuario.perfil.in_(['Equipe','Técnico'])).first():
        db.session.add(Usuario(nome='Equipe BR',email='equipe@brsolucoes.com.br',senha=ph('123456'),perfil='Equipe',ativo=True))
    db.session.commit()
    if not Chamado.query.first():
        p=BASE/'data'/'seed.json'
        if p.exists():
            data=json.loads(p.read_text(encoding='utf-8'))
            maps=[('empreendimentos',Empreendimento),('chamados',Chamado),('garantias',Garantia),('fotos',Foto),('materiais',Material),('financeiro',Financeiro)]
            for key,model in maps:
                cols={c.name for c in model.__table__.columns}
                for row in data.get(key,[]):
                    row={k:v for k,v in row.items() if k in cols and k!='id'}
                    db.session.add(model(**row))
            db.session.commit()

def login_required(fn):
    @wraps(fn)
    def w(*a,**k):
        if not session.get('user_id'): return redirect(url_for('login'))
        return fn(*a,**k)
    return w

def allow(*profiles):
    def deco(fn):
        @wraps(fn)
        def w(*a,**k):
            if session.get('perfil') not in profiles:
                flash('Seu usuário não possui permissão para acessar esta área.','erro')
                return redirect(url_for('dashboard'))
            return fn(*a,**k)
        return w
    return deco

def can_edit(): return session.get('perfil') in ('Administrador','Líder','Equipe','Técnico')
def is_admin(): return session.get('perfil')=='Administrador'

def agenda_destino_filter(tipo):
    nome=(session.get('nome') or '').strip().lower()
    email=(session.get('email') or '').strip().lower()
    uid=session.get('user_id')
    if tipo=='lider':
        filtros=[Agenda.lider_id==uid]
        if nome: filtros.append(func.lower(Agenda.lider)==nome)
        if email: filtros.append(func.lower(Agenda.lider_email)==email)
    else:
        filtros=[Agenda.equipe_id==uid]
        if nome: filtros.append(func.lower(Agenda.equipe)==nome)
        if email: filtros.append(func.lower(Agenda.equipe_email)==email)
    return or_(*filtros)

def log(cid,acao):
    db.session.add(Historico(chamado_id=cid,usuario=session.get('nome','Sistema'),acao=acao)); db.session.commit()

def audit(modulo,registro_id,acao,detalhes=''):
    db.session.add(Auditoria(modulo=modulo,registro_id=registro_id,usuario=session.get('nome','Sistema'),acao=acao,detalhes=detalhes)); db.session.commit()

def snapshot(obj, fields):
    return ' | '.join(f'{f}: {getattr(obj,f,None)}' for f in fields)

def guess_category(text):
    import re
    t=(text or '').lower(); rules=[
      ('infiltra|vazamento|umidade|mofo','Instalações hidráulicas - vedação / vazamento'),('janela|esquadria|vedação','Esquadrias de alumínio - vedação e funcionamento'),('telha|telhado|cobertura|calha','Cobertura e telhados'),('drywall|gesso','Drywall - fissuras'),('pintura|descasc|empol|tinta','Pintura interna'),('fissura|trinca|rachadura','Revestimento em argamassa - fissuras'),('estrutura|concreto','Estrutura principal - solidez e segurança')]
    for pat,cat in rules:
        if re.search(pat,t): return cat
    return 'Outros / análise técnica'

def warranty_analysis(emp,cat,de,da):
    row=Garantia.query.filter(Garantia.item==cat,Garantia.empreendimento.in_([emp,'Todos'])).order_by(case((Garantia.empreendimento==emp,0),else_=1)).first()
    if not row: return {'resultado':'Necessita vistoria','prazo':'','fundamentacao':'Item não localizado na base resumida. Consulte o manual completo e realize vistoria técnica.'}
    elapsed=None
    try:
        d1=datetime.strptime(de,'%Y-%m-%d').date(); d2=datetime.strptime(da,'%Y-%m-%d').date(); elapsed=(d2.year-d1.year)*12+d2.month-d1.month-(1 if d2.day<d1.day else 0)
    except: pass
    if row.prazo_meses==0: result='Improcedente'
    elif elapsed is None: result='Necessita vistoria'
    elif elapsed<=row.prazo_meses: result=row.classificacao_padrao or 'Necessita vistoria'
    else: result='Fora do prazo / analisar improcedência'
    return {'resultado':result,'prazo':f'{row.prazo_meses} meses' if row.prazo_meses else 'No ato da entrega','fundamentacao':row.descricao,'fonte':row.fonte,'meses_decorridos':elapsed}

@app.context_processor
def context():
    avisos_lider=0
    if session.get('perfil')=='Líder' and session.get('nome'):
        avisos_lider=Agenda.query.filter(agenda_destino_filter('lider'),Agenda.lider_confirmou==False,Agenda.status.in_(['Aguardando Líder','Aguardando agendamento do líder'])).count()
    return dict(can_edit=can_edit(),is_admin=is_admin(),perfil=session.get('perfil'),now=datetime.now(),avisos_lider=avisos_lider,app_version=APP_VERSION,agenda_status=['Aguardando Líder','Agendada pelo Administrador','Agendada pelo Líder','Enviada para equipe','Em Execução','Finalizada','Cliente não respondeu','Reagendada','Cancelada','Atendimento realizado'])

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        u=Usuario.query.filter(func.lower(Usuario.email)==request.form.get('email','').strip().lower(),Usuario.ativo==True).first()
        if u and check(u.senha,request.form.get('senha','')):
            session.clear(); session.update(user_id=u.id,nome=u.nome,email=u.email,perfil=u.perfil,construtora=u.construtora); return redirect(url_for('dashboard'))
        flash('E-mail ou senha inválidos.','erro')
    return render_template('login.html')
@app.route('/sair')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    base_q=Chamado.query
    if session.get('perfil')=='Construtora': base_q=base_q.filter(Chamado.construtora==session.get('construtora'))
    total=base_q.count(); final=base_q.filter(func.lower(Chamado.status).like('%final%')).count(); impro=base_q.filter(func.lower(Chamado.classificacao).like('%improced%')).count(); abertos=total-final
    por_emp=db.session.query(Chamado.empreendimento,func.count(Chamado.id))
    if session.get('perfil')=='Construtora': por_emp=por_emp.filter(Chamado.construtora==session.get('construtora'))
    por_emp=por_emp.group_by(Chamado.empreendimento).order_by(func.count(Chamado.id).desc()).all()
    recentes=base_q.order_by(Chamado.id.desc()).limit(8).all()
    fin=None
    if is_admin():
        entradas=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter(Financeiro.tipo=='Entrada',Financeiro.status=='Pago/Recebido').scalar(); saidas=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter(Financeiro.tipo=='Saída',Financeiro.status=='Pago/Recebido').scalar(); fin={'entradas':entradas,'saidas':saidas}
    em_campo=Atendimento.query.filter_by(tipo='checkin').count()-Atendimento.query.filter_by(tipo='checkout').count()
    status_chart=db.session.query(Chamado.status,func.count(Chamado.id)).group_by(Chamado.status).order_by(func.count(Chamado.id).desc()).all()
    emp_chart=[(x[0] or 'Sem empreendimento',x[1]) for x in por_emp[:8]]
    agenda_hoje=Agenda.query.filter(Agenda.data_agendada==date.today().isoformat())
    if session.get('perfil')=='Construtora': agenda_hoje=agenda_hoje.filter(Agenda.construtora==session.get('construtora'))
    agenda_hoje=agenda_hoje.count()
    return render_template('dashboard.html',k={'total':total,'finalizados':final,'abertos':abertos,'improcedentes':impro,'em_campo':max(em_campo,0),'agenda_hoje':agenda_hoje},por_emp=por_emp,recentes=recentes,fin=fin,status_chart=status_chart,emp_chart=emp_chart)

@app.route('/chamados')
@login_required
def chamados():
    q=request.args.get('q','').strip(); emp=request.args.get('empreendimento',''); status=request.args.get('status',''); qry=Chamado.query
    if session.get('perfil')=='Construtora': qry=qry.filter(Chamado.construtora==session.get('construtora'))
    if q: qry=qry.filter(or_(Chamado.sat.ilike(f'%{q}%'),Chamado.solicitante.ilike(f'%{q}%'),Chamado.unidade.ilike(f'%{q}%'),Chamado.problema.ilike(f'%{q}%')))
    if emp: qry=qry.filter_by(empreendimento=emp)
    if status: qry=qry.filter_by(status=status)
    return render_template('chamados.html',rows=qry.order_by(Chamado.id.desc()).all(),emps=Empreendimento.query.order_by(Empreendimento.nome).all(),statuses=[x[0] for x in db.session.query(Chamado.status).distinct().order_by(Chamado.status)])

@app.route('/chamados/novo',methods=['GET','POST'])
@login_required
@allow('Administrador','Líder','Equipe')
def novo_chamado():
    if request.method=='POST':
        cat=request.form.get('categoria') or guess_category(request.form.get('problema')); da=request.form.get('data_recebido') or date.today().isoformat(); ana=warranty_analysis(request.form.get('empreendimento'),cat,request.form.get('data_entrega'),da)
        c=Chamado(sat=request.form.get('sat'),empreendimento=request.form.get('empreendimento'),data_recebido=da,solicitante=request.form.get('solicitante'),unidade=request.form.get('unidade'),problema=request.form.get('problema'),contato=request.form.get('contato'),status=request.form.get('status') or 'Aberto',classificacao=request.form.get('classificacao') or ana['resultado'],categoria=cat,data_entrega=request.form.get('data_entrega'),analise_garantia=ana['resultado'],fundamentacao=ana['fundamentacao'],responsavel=request.form.get('responsavel'),observacoes=request.form.get('observacoes'))
        db.session.add(c); db.session.commit(); log(c.id,'Chamado cadastrado'); flash('Chamado cadastrado com sucesso.','ok'); return redirect(url_for('ver_chamado',id=c.id))
    return render_template('chamado_form.html',emps=Empreendimento.query.order_by(Empreendimento.nome).all(),cats=[x[0] for x in db.session.query(Garantia.item).distinct().order_by(Garantia.item)])


@app.route('/chamados/importar',methods=['GET','POST'])
@login_required
@allow('Administrador','Líder','Equipe')
def importar_chamados():
    if request.method=='POST':
        f=request.files.get('planilha')
        if not f or not f.filename:
            flash('Selecione uma planilha.','erro'); return redirect(url_for('importar_chamados'))
        try:
            rows,mapping=read_import_file(f)
            sats=[r.get('sat','').strip() for r in rows if r.get('sat')]
            existing={x[0] for x in db.session.query(Chamado.sat).filter(Chamado.sat.in_(sats)).all() if x[0]} if sats else set()
            seen=set()
            for r in rows:
                sat=r.get('sat','').strip()
                if sat and (sat in existing or sat in seen): r['_duplicada']=True
                else: r['_duplicada']=False
                if sat: seen.add(sat)
            token=uuid.uuid4().hex
            (IMPORT_DIR/f'{token}.json').write_text(json.dumps(rows,ensure_ascii=False),encoding='utf-8')
            return render_template('importar_chamados.html',rows=rows,token=token,mapping=mapping,preview=True)
        except Exception as e:
            flash(str(e),'erro')
    return render_template('importar_chamados.html',preview=False)

@app.post('/chamados/importar/confirmar')
@login_required
@allow('Administrador','Líder','Equipe')
def confirmar_importacao():
    token=re.sub(r'[^a-f0-9]','',request.form.get('token',''))
    path=IMPORT_DIR/f'{token}.json'
    if not path.exists():
        flash('A prévia expirou. Envie a planilha novamente.','erro'); return redirect(url_for('importar_chamados'))
    rows=json.loads(path.read_text(encoding='utf-8')); imported=duplicates=errors=0
    for r in rows:
        if r.get('_erro'): errors+=1; continue
        sat=(r.get('sat') or '').strip()
        if sat and Chamado.query.filter_by(sat=sat).first(): duplicates+=1; continue
        emp=(r.get('empreendimento') or '').strip()
        if emp and not Empreendimento.query.filter_by(nome=emp).first(): db.session.add(Empreendimento(nome=emp))
        cat=r.get('categoria') or guess_category(r.get('problema'))
        da=r.get('data_recebido') or date.today().isoformat()
        ana=warranty_analysis(emp,cat,r.get('data_entrega'),da)
        c=Chamado(sat=sat,empreendimento=emp,data_recebido=da,solicitante=r.get('solicitante'),unidade=r.get('unidade'),problema=r.get('problema'),contato=r.get('contato'),status=r.get('status') or 'Aberto',classificacao=r.get('classificacao') or ana['resultado'],categoria=cat,data_entrega=r.get('data_entrega'),analise_garantia=ana['resultado'],fundamentacao=ana['fundamentacao'],responsavel=r.get('responsavel'),observacoes=r.get('observacoes'))
        db.session.add(c); db.session.flush(); db.session.add(Historico(chamado_id=c.id,usuario=session.get('nome','Sistema'),acao='Chamado importado por planilha'))
        imported+=1
    db.session.commit()
    try: path.unlink()
    except OSError: pass
    flash(f'Importação concluída: {imported} chamados incluídos, {duplicates} duplicados ignorados e {errors} linhas com erro.','ok')
    return redirect(url_for('chamados'))

@app.route('/chamados/modelo-importacao.xlsx')
@login_required
def modelo_importacao():
    wb=Workbook(); ws=wb.active; ws.title='Chamados'
    headers=['SAT','Empreendimento','Data de abertura','Morador','Unidade','Problema','Contato','Status','Classificação','Categoria','Data de entrega','Responsável','Observações']
    ws.append(headers); ws.append(['26576','Vittace Reserva','20/07/2026','Maria da Silva','Torre 06 - Apto 403','Infiltração no teto do banheiro','(43) 99999-9999','Aberto','','','15/03/2023','',''])
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:M2'
    widths=[14,24,18,24,22,48,20,16,20,28,18,22,38]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name='Modelo_Importacao_BR_SmartSAT.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/chamados/<int:id>',methods=['GET','POST'])
@login_required
def ver_chamado(id):
    c=Chamado.query.get_or_404(id)
    if request.method=='POST':
        if not can_edit(): abort(403)
        antes=snapshot(c,['sat','empreendimento','data_recebido','solicitante','unidade','problema','contato','status','classificacao','categoria','data_entrega','responsavel','observacoes'])
        for field in ['sat','empreendimento','data_recebido','solicitante','unidade','problema','contato','status','classificacao','categoria','data_entrega','responsavel','observacoes']:
            if field in request.form: setattr(c,field,request.form.get(field))
        db.session.commit(); log(id,f"Chamado atualizado para status: {c.status}"); audit('Chamado',id,'EDITADO',antes+' => '+snapshot(c,['sat','empreendimento','data_recebido','solicitante','unidade','problema','contato','status','classificacao','categoria','data_entrega','responsavel','observacoes'])); flash('Chamado atualizado.','ok'); return redirect(url_for('ver_chamado',id=id))
    drive_folder_link=None
    drive_sync_message=None
    if drive_configured():
        try:
            # Ao abrir a SAT, procura a pasta correspondente no Drive e vincula
            # automaticamente as fotos ainda não registradas no sistema.
            result=sync_drive_photos(c)
            if result:
                drive_folder_link=(result.get('pasta') or {}).get('webViewLink')
                if result.get('importadas'):
                    drive_sync_message=f"{result['importadas']} foto(s) nova(s) sincronizada(s) do Google Drive."
            if not drive_folder_link:
                existing=Foto.query.filter_by(chamado_id=id).filter(Foto.drive_folder_link.isnot(None)).first()
                drive_folder_link=existing.drive_folder_link if existing else None
        except Exception as exc:
            app.logger.exception('Falha na sincronização automática do Google Drive')
            drive_sync_message='Não foi possível consultar o Google Drive agora. As fotos já salvas continuam disponíveis.'
    fotos=Foto.query.filter_by(chamado_id=id).order_by(Foto.id.desc()).all()
    grupos_fotos={
        'Antes do serviço':[],
        'Durante o serviço':[],
        'Depois do serviço':[],
        'Vistoria':[],
        'Material utilizado':[],
        'Outras fotos':[]
    }
    for foto_item in fotos:
        chave=foto_item.descricao if foto_item.descricao in grupos_fotos else 'Outras fotos'
        grupos_fotos[chave].append(foto_item)
    return render_template('chamado_detalhe.html',c=c,fotos=fotos,grupos_fotos=grupos_fotos,drive_folder_link=drive_folder_link if api_drive_configured() else None,drive_local_folder=(result.get('pasta') or {}).get('name') if drive_configured() and 'result' in locals() and result.get('pasta') else None,drive_sync_message=drive_sync_message,assinaturas=Assinatura.query.filter_by(chamado_id=id).order_by(Assinatura.id.desc()).all(),mats=Material.query.filter_by(chamado_id=id).order_by(Material.id.desc()).all(),hist=Historico.query.filter_by(chamado_id=id).order_by(Historico.id.desc()).limit(50).all(),atendimentos=Atendimento.query.filter_by(chamado_id=id).order_by(Atendimento.id.desc()).all(),relatorio=RelatorioTecnico.query.filter_by(chamado_id=id).first(),drive_ativo=drive_configured(),drive_modo=drive_mode(),cats=[x[0] for x in db.session.query(Garantia.item).distinct().order_by(Garantia.item)])


@app.post('/chamados/<int:id>/excluir')
@login_required
@allow('Administrador','Líder')
def excluir_chamado(id):
    c=Chamado.query.get_or_404(id)
    detalhes=snapshot(c,['sat','empreendimento','solicitante','unidade','status'])
    excluir_drive=request.form.get('excluir_drive')=='1'
    # Remove arquivos físicos e registros vinculados. Os arquivos do Google Drive
    # só são apagados quando o usuário marca essa opção na confirmação.
    for f in Foto.query.filter_by(chamado_id=id).all():
        if excluir_drive and f.drive_file_id:
            delete_drive_file(f.drive_file_id)
        if f.arquivo:
            try: (UPLOAD/f.arquivo).unlink(missing_ok=True)
            except Exception: pass
        db.session.delete(f)
    for a in Assinatura.query.filter_by(chamado_id=id).all():
        try: (UPLOAD/a.arquivo).unlink(missing_ok=True)
        except Exception: pass
        db.session.delete(a)
    Material.query.filter_by(chamado_id=id).delete(synchronize_session=False)
    Agenda.query.filter_by(chamado_id=id).delete(synchronize_session=False)
    Atendimento.query.filter_by(chamado_id=id).delete(synchronize_session=False)
    RelatorioTecnico.query.filter_by(chamado_id=id).delete(synchronize_session=False)
    Historico.query.filter_by(chamado_id=id).delete(synchronize_session=False)
    db.session.delete(c); db.session.commit(); audit('Chamado',id,'EXCLUÍDO',detalhes)
    flash('SAT excluída. '+('As fotos também foram removidas do Google Drive.' if excluir_drive else 'As fotos foram mantidas no Google Drive.'),'ok'); return redirect(url_for('chamados'))


@app.route('/configurar-drive',methods=['GET','POST'])
@login_required
@allow('Administrador')
def configurar_drive():
    atual=str(local_drive_root() or '')
    teste=None
    if request.method=='POST':
        raw=(request.form.get('local_root') or '').strip().strip('"')
        path=Path(raw).expanduser()
        if not path.exists() or not path.is_dir():
            flash('A pasta informada não existe. Copie o caminho completo da pasta do Google Drive.','erro')
        else:
            DRIVE_CONFIG_FILE.write_text(json.dumps({'local_root':str(path.resolve())},ensure_ascii=False,indent=2),encoding='utf-8')
            flash('Google Drive para computador configurado com sucesso.','ok')
            return redirect(url_for('configurar_drive'))
    if local_drive_root():
        try:
            count=sum(1 for current,dirs,files in os.walk(local_drive_root()))
            teste=f'Pasta acessível. {count} pasta(s) foram identificadas para pesquisa.'
        except Exception as exc:
            teste='Não foi possível ler a pasta: '+str(exc)
    return render_template('configurar_drive.html',atual=atual,modo=drive_mode(),teste=teste)

@app.post('/configurar-drive/desativar')
@login_required
@allow('Administrador')
def desativar_drive_local():
    try: DRIVE_CONFIG_FILE.unlink(missing_ok=True)
    except Exception: pass
    flash('Integração local do Google Drive desativada.','ok')
    return redirect(url_for('configurar_drive'))

@app.post('/chamados/<int:id>/sincronizar-drive')
@login_required
@allow('Administrador','Líder','Equipe')
def sincronizar_drive(id):
    chamado=Chamado.query.get_or_404(id)
    if not drive_configured():
        flash('Configure o Google Drive para sincronizar as fotos existentes.','erro')
        return redirect(url_for('ver_chamado',id=id))
    try:
        result=sync_drive_photos(chamado)
        if result['pasta'] is None:
            flash('A pasta desta SAT não foi encontrada no Drive. Verifique o padrão Empreendimento / SAT número.','erro')
        elif result['importadas']:
            flash(f"{result['importadas']} foto(s) do Google Drive foram vinculadas à SAT.",'ok')
        else:
            flash('Sincronização concluída. Todas as fotos do Drive já estavam vinculadas à SAT.','ok')
    except Exception as exc:
        app.logger.exception('Falha ao sincronizar fotos do Drive')
        flash('Não foi possível sincronizar o Google Drive: '+str(exc),'erro')
    return redirect(url_for('ver_chamado',id=id))

@app.get('/fotos/<int:id>/arquivo-drive')
@login_required
def arquivo_drive(id):
    foto=Foto.query.get_or_404(id)
    if not foto.drive_file_id or not drive_configured():
        abort(404)
    if str(foto.drive_file_id).startswith('local:'):
        sat=Chamado.query.get_or_404(foto.chamado_id)
        folder=local_drive_locate_sat_folder(sat)
        if not folder: abort(404)
        token=foto.drive_file_id
        for path in local_drive_image_files(folder):
            candidate='local:'+hashlib.sha256(str(path.resolve()).encode('utf-8')).hexdigest()
            if candidate==token:
                return send_file(path,max_age=300)
        abort(404)
    service=drive_service()
    metadata=service.files().get(fileId=foto.drive_file_id,fields='name,mimeType',supportsAllDrives=True).execute()
    request_media=service.files().get_media(fileId=foto.drive_file_id,supportsAllDrives=True)
    buffer=io.BytesIO()
    downloader=MediaIoBaseDownload(buffer,request_media)
    done=False
    while not done:
        _,done=downloader.next_chunk()
    buffer.seek(0)
    return send_file(buffer,mimetype=metadata.get('mimeType') or 'image/jpeg',download_name=metadata.get('name') or 'foto.jpg',max_age=3600)

@app.post('/fotos/<int:id>/editar')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def editar_foto(id):
    f=Foto.query.get_or_404(id); antes=f.descricao or ''
    f.descricao=request.form.get('descricao') or f.descricao; db.session.commit(); log(f.chamado_id,f'Foto editada: {antes} para {f.descricao}'); audit('Foto',id,'EDITADA',antes+' => '+(f.descricao or ''))
    flash('Descrição da foto atualizada.','ok'); return redirect(url_for('ver_chamado',id=f.chamado_id))

@app.post('/fotos/<int:id>/excluir')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def excluir_foto(id):
    f=Foto.query.get_or_404(id); cid=f.chamado_id; detalhes=f'{f.descricao} | {f.arquivo}'
    delete_drive_file(f.drive_file_id)
    try: (UPLOAD/f.arquivo).unlink(missing_ok=True)
    except Exception: pass
    db.session.delete(f); db.session.commit(); log(cid,'Foto excluída'); audit('Foto',id,'EXCLUÍDA',detalhes)
    flash('Foto excluída.','ok'); return redirect(url_for('ver_chamado',id=cid))

@app.post('/assinaturas/<int:id>/editar')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def editar_assinatura(id):
    a=Assinatura.query.get_or_404(id); antes=snapshot(a,['nome_cliente','documento','observacao'])
    a.nome_cliente=request.form.get('nome_cliente') or a.nome_cliente; a.documento=request.form.get('documento'); a.observacao=request.form.get('observacao'); db.session.commit(); log(a.chamado_id,'Dados da assinatura atualizados'); audit('Assinatura',id,'EDITADA',antes+' => '+snapshot(a,['nome_cliente','documento','observacao']))
    flash('Dados da assinatura atualizados.','ok'); return redirect(url_for('ver_chamado',id=a.chamado_id))

@app.post('/assinaturas/<int:id>/excluir')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def excluir_assinatura(id):
    a=Assinatura.query.get_or_404(id); cid=a.chamado_id; detalhes=snapshot(a,['nome_cliente','documento','assinado_em'])
    try: (UPLOAD/a.arquivo).unlink(missing_ok=True)
    except Exception: pass
    db.session.delete(a); db.session.commit(); log(cid,'Assinatura excluída para nova coleta'); audit('Assinatura',id,'EXCLUÍDA',detalhes)
    flash('Assinatura excluída. Uma nova assinatura já pode ser coletada.','ok'); return redirect(url_for('ver_chamado',id=cid))

@app.post('/materiais/<int:id>/editar')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def editar_material(id):
    m=Material.query.get_or_404(id); antes=snapshot(m,['item','quantidade','status','observacoes'])
    for field in ['item','quantidade','status','observacoes']: setattr(m,field,request.form.get(field))
    db.session.commit(); log(m.chamado_id,'Material editado'); audit('Material',id,'EDITADO',antes+' => '+snapshot(m,['item','quantidade','status','observacoes']))
    flash('Material atualizado.','ok'); return redirect(url_for('ver_chamado',id=m.chamado_id))

@app.post('/materiais/<int:id>/excluir')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def excluir_material(id):
    m=Material.query.get_or_404(id); cid=m.chamado_id; detalhes=snapshot(m,['item','quantidade','status'])
    db.session.delete(m); db.session.commit(); log(cid,'Material excluído'); audit('Material',id,'EXCLUÍDO',detalhes)
    flash('Material excluído.','ok'); return redirect(url_for('ver_chamado',id=cid))

@app.post('/chamados/<int:id>/analisar')
@login_required
@allow('Administrador','Líder','Equipe')
def analisar_chamado(id):
    c=Chamado.query.get_or_404(id); cat=request.form.get('categoria') or guess_category(c.problema); a=warranty_analysis(c.empreendimento,cat,c.data_entrega,c.data_recebido); c.categoria=cat;c.analise_garantia=a['resultado'];c.fundamentacao=a['fundamentacao']+((' Fonte: '+a.get('fonte','')) if a.get('fonte') else '');db.session.commit();log(id,'Garantia analisada');return redirect(url_for('ver_chamado',id=id))
@app.post('/chamados/<int:id>/foto')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def foto(id):
    Chamado.query.get_or_404(id)
    f=request.files.get('foto')
    if not f or not f.filename:
        flash('Selecione ou tire uma foto.','erro'); return redirect(url_for('ver_chamado',id=id))
    loc=parse_location(request.form)
    if not loc:
        flash('A localização é obrigatória. Ative o GPS e permita o acesso à localização.','erro'); return redirect(url_for('ver_chamado',id=id))
    lat,lon,acc=loc
    chamado=Chamado.query.get_or_404(id)
    original=secure_filename(f.filename) or 'foto.jpg'
    name=f'{uuid.uuid4().hex}_{original}'
    local_path=UPLOAD/name
    f.save(local_path)
    etapa=request.form.get('descricao') or 'Fotos'
    drive_data=None
    drive_error=None
    if drive_configured():
        try:
            drive_data=upload_photo_to_drive(local_path,chamado,etapa,original)
        except Exception as exc:
            drive_error=str(exc)
            app.logger.exception('Falha ao enviar foto ao Google Drive')
    foto_reg=Foto(chamado_id=id,arquivo=name,descricao=etapa,latitude=lat,longitude=lon,precisao=acc,capturado_em=parse_client_time(request.form.get('capturado_em')),usuario=session.get('nome'),dispositivo=(request.form.get('dispositivo') or '')[:255],drive_file_id=(drive_data or {}).get('file_id'),drive_web_link=(drive_data or {}).get('web_link'),drive_folder_link=(drive_data or {}).get('folder_link'),armazenamento='Google Drive' if drive_data else 'Local')
    db.session.add(foto_reg)
    db.session.commit();log(id,f'Foto adicionada com GPS: {lat:.6f}, {lon:.6f} ({foto_reg.armazenamento})')
    if drive_data:
        flash('Foto salva na SAT e enviada ao Google Drive.','ok')
    elif drive_error:
        flash('Foto salva localmente, mas o envio ao Google Drive falhou: '+drive_error,'erro')
    else:
        flash('Foto e localização registradas localmente. Configure o Google Drive para envio automático.','ok')
    return redirect(url_for('ver_chamado',id=id))

@app.post('/chamados/<int:id>/assinatura')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def assinatura(id):
    Chamado.query.get_or_404(id)
    data=request.form.get('assinatura_data','')
    if not data.startswith('data:image/png;base64,'):
        flash('Faça a assinatura no quadro antes de salvar.','erro'); return redirect(url_for('ver_chamado',id=id))
    try:
        raw=base64.b64decode(data.split(',',1)[1])
        name=f'assinatura_{id}_{uuid.uuid4().hex}.png'
        (UPLOAD/name).write_bytes(raw)
        loc=parse_location(request.form)
        if not loc:
            flash('A localização é obrigatória para registrar a assinatura. Ative o GPS e tente novamente.','erro'); return redirect(url_for('ver_chamado',id=id))
        lat,lon,acc=loc
        db.session.add(Assinatura(chamado_id=id,arquivo=name,nome_cliente=request.form.get('nome_cliente'),documento=request.form.get('documento'),observacao=request.form.get('observacao'),latitude=lat,longitude=lon,precisao=acc,assinado_em=parse_client_time(request.form.get('assinado_em')),usuario=session.get('nome'),dispositivo=(request.form.get('dispositivo') or '')[:255]))
        db.session.commit(); log(id,f"Assinatura do cliente registrada com GPS: {lat:.6f}, {lon:.6f}")
        flash('Assinatura registrada com sucesso.','ok')
    except Exception:
        flash('Não foi possível salvar a assinatura. Tente novamente.','erro')
    return redirect(url_for('ver_chamado',id=id))

@app.post('/chamados/<int:id>/material')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def material(id):
    db.session.add(Material(chamado_id=id,item=request.form.get('item'),quantidade=request.form.get('quantidade'),status=request.form.get('status'),observacoes=request.form.get('observacoes')));db.session.commit();log(id,'Material registrado');return redirect(url_for('ver_chamado',id=id))


@app.post('/chamados/<int:id>/atendimento/<tipo>')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def registrar_atendimento(id,tipo):
    Chamado.query.get_or_404(id)
    if tipo not in ('checkin','checkout'): abort(400)
    loc=parse_location(request.form)
    if not loc:
        flash('A localização é obrigatória para registrar chegada ou saída.','erro'); return redirect(url_for('ver_chamado',id=id))
    lat,lon,acc=loc
    db.session.add(Atendimento(chamado_id=id,tipo=tipo,latitude=lat,longitude=lon,precisao=acc,usuario=session.get('nome'),dispositivo=(request.form.get('dispositivo') or '')[:255],observacao=request.form.get('observacao')))
    db.session.commit(); log(id,'Check-in registrado' if tipo=='checkin' else 'Check-out registrado')
    flash(('Chegada' if tipo=='checkin' else 'Saída')+' registrada com GPS.','ok'); return redirect(url_for('ver_chamado',id=id))

@app.post('/chamados/<int:id>/relatorio')
@login_required
@allow('Administrador','Líder','Equipe','Técnico')
def salvar_relatorio(id):
    Chamado.query.get_or_404(id)
    r=RelatorioTecnico.query.filter_by(chamado_id=id).first() or RelatorioTecnico(chamado_id=id)
    r.diagnostico=request.form.get('diagnostico'); r.servico_executado=request.form.get('servico_executado'); r.testes_realizados=request.form.get('testes_realizados'); r.conclusao=request.form.get('conclusao'); r.usuario=session.get('nome')
    db.session.add(r); db.session.commit(); log(id,'Relatório técnico atualizado'); flash('Relatório técnico salvo.','ok'); return redirect(url_for('ver_chamado',id=id))

@app.route('/chamados/<int:id>/pdf')
@login_required
def pdf_chamado(id):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    c=Chamado.query.get_or_404(id); rel=RelatorioTecnico.query.filter_by(chamado_id=id).first(); ass=Assinatura.query.filter_by(chamado_id=id).order_by(Assinatura.id.desc()).first(); fotos=Foto.query.filter_by(chamado_id=id).order_by(Foto.id).all()
    buf=io.BytesIO(); pdf=canvas.Canvas(buf,pagesize=A4); W,H=A4; y=H-45
    def line(txt,bold=False,size=10):
        nonlocal y
        pdf.setFont('Helvetica-Bold' if bold else 'Helvetica',size)
        for part in [txt[i:i+95] for i in range(0,len(txt or ''),95)] or ['']:
            if y<60: pdf.showPage(); y=H-45
            pdf.drawString(40,y,part); y-=14
    line('BR SMARTSAT ENTERPRISE 12.0 - RELATORIO DE ATENDIMENTO',True,14); line(f'SAT: {c.sat or c.id}',True); line(f'Empreendimento: {c.empreendimento}'); line(f'Unidade: {c.unidade}'); line(f'Morador: {c.solicitante}'); line(f'Status: {c.status} | Classificacao: {c.classificacao}'); line('Problema informado:',True); line(c.problema or '-')
    if rel:
        line('Diagnostico tecnico:',True); line(rel.diagnostico or '-'); line('Servico executado:',True); line(rel.servico_executado or '-'); line('Testes realizados:',True); line(rel.testes_realizados or '-'); line('Conclusao:',True); line(rel.conclusao or '-')
    line(f'Fotos registradas: {len(fotos)}',True)
    for f in fotos[:6]:
        path=UPLOAD/f.arquivo
        if path.exists():
            if y<180: pdf.showPage(); y=H-45
            try: pdf.drawImage(ImageReader(str(path)),40,y-130,width=180,height=125,preserveAspectRatio=True,anchor='sw'); pdf.drawString(230,y-20,(f.descricao or 'Foto')[:50]); pdf.drawString(230,y-38,f'GPS: {f.latitude}, {f.longitude}'); y-=145
            except: pass
    if ass:
        line('Assinatura do cliente:',True); line(f'{ass.nome_cliente or "Cliente"} - {ass.documento or "documento nao informado"}')
        path=UPLOAD/ass.arquivo
        if path.exists():
            if y<160: pdf.showPage(); y=H-45
            try: pdf.drawImage(ImageReader(str(path)),40,y-110,width=300,height=100,preserveAspectRatio=True,anchor='sw'); y-=120
            except: pass
    line(f'Codigo de validacao: SAT-{c.id}-{(c.sat or c.id)}',True); pdf.save(); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f'SAT_{c.sat or c.id}_BR_SmartSAT.pdf',mimetype='application/pdf')

@app.route('/validar/<int:id>')
def validar(id):
    c=Chamado.query.get_or_404(id); ass=Assinatura.query.filter_by(chamado_id=id).order_by(Assinatura.id.desc()).first(); return render_template('validar.html',c=c,ass=ass)

@app.route('/manifest.webmanifest')
def manifest(): return send_file(BASE/'static'/'manifest.webmanifest',mimetype='application/manifest+json')
@app.route('/service-worker.js')
def sw(): return send_file(BASE/'static'/'service-worker.js',mimetype='application/javascript')

@app.route('/garantias')
@login_required
def garantias():
    emp=request.args.get('empreendimento','');q=request.args.get('q','');qry=Garantia.query
    if emp: qry=qry.filter(Garantia.empreendimento.in_([emp,'Todos']))
    if q: qry=qry.filter(or_(Garantia.item.ilike(f'%{q}%'),Garantia.descricao.ilike(f'%{q}%')))
    return render_template('garantias.html',rows=qry.order_by(Garantia.empreendimento,Garantia.item,Garantia.prazo_meses).all())

@app.route('/financeiro',methods=['GET','POST'])
@login_required
@allow('Administrador')
def financeiro():
    if request.method=='POST':
        r=Financeiro(data=request.form.get('data'),tipo=request.form.get('tipo'),descricao=request.form.get('descricao'),origem=request.form.get('origem'),construtora=request.form.get('construtora'),empreendimento=request.form.get('empreendimento'),categoria=request.form.get('categoria'),valor=float(request.form.get('valor') or 0),status=request.form.get('status'),forma_pagamento=request.form.get('forma_pagamento'),vencimento=request.form.get('vencimento'),documento=request.form.get('documento'),observacoes=request.form.get('observacoes'))
        db.session.add(r);db.session.commit();audit('Financeiro',r.id,'CRIADO',snapshot(r,['data','tipo','descricao','origem','construtora','empreendimento','valor','status']));flash('Lançamento financeiro salvo.','ok');return redirect(url_for('financeiro'))
    rows=Financeiro.query.order_by(Financeiro.data.desc(),Financeiro.id.desc()).all()
    ent=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter_by(tipo='Entrada',status='Pago/Recebido').scalar();sai=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter_by(tipo='Saída',status='Pago/Recebido').scalar()
    pend_ent=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter(Financeiro.tipo=='Entrada',Financeiro.status!='Pago/Recebido').scalar(); pend_sai=db.session.query(func.coalesce(func.sum(Financeiro.valor),0)).filter(Financeiro.tipo=='Saída',Financeiro.status!='Pago/Recebido').scalar()
    mensal={}
    for r in rows:
        mes=(r.data or '')[:7] or 'Sem data'
        mensal.setdefault(mes,{'entrada':0.0,'saida':0.0})
        if r.status=='Pago/Recebido':
            mensal[mes]['entrada' if r.tipo=='Entrada' else 'saida']+=r.valor or 0
    chart_meses=sorted(mensal)[-12:]
    chart_data=[{'mes':m,'entrada':mensal[m]['entrada'],'saida':mensal[m]['saida']} for m in chart_meses]
    por_origem={}
    for r in rows:
        if r.tipo=='Entrada' and r.status=='Pago/Recebido': por_origem[r.origem or r.construtora or r.empreendimento or 'Outras'] = por_origem.get(r.origem or r.construtora or r.empreendimento or 'Outras',0)+(r.valor or 0)
    origem_chart=sorted(por_origem.items(),key=lambda x:x[1],reverse=True)[:8]
    return render_template('financeiro.html',rows=rows,sums={'entradas':ent,'saidas':sai,'saldo':ent-sai,'receber':pend_ent,'pagar':pend_sai},emps=Empreendimento.query.order_by(Empreendimento.nome).all(),construtoras=Construtora.query.order_by(Construtora.nome).all(),chart_data=chart_data,origem_chart=origem_chart)

@app.route('/financeiro/<int:id>/editar',methods=['GET','POST'])
@login_required
@allow('Administrador')
def editar_financeiro(id):
    r=Financeiro.query.get_or_404(id)
    if request.method=='POST':
        antes=snapshot(r,['data','tipo','descricao','origem','construtora','empreendimento','valor','status'])
        for f in ['data','tipo','descricao','origem','construtora','empreendimento','categoria','status','forma_pagamento','vencimento','documento','observacoes']: setattr(r,f,request.form.get(f))
        r.valor=float(request.form.get('valor') or 0); db.session.commit(); audit('Financeiro',id,'EDITADO',antes); flash('Lançamento atualizado.','ok'); return redirect(url_for('financeiro'))
    return render_template('financeiro_editar.html',r=r,emps=Empreendimento.query.order_by(Empreendimento.nome).all(),construtoras=Construtora.query.order_by(Construtora.nome).all())

@app.post('/financeiro/<int:id>/excluir')
@login_required
@allow('Administrador')
def excluir_financeiro(id):
    r=Financeiro.query.get_or_404(id); db.session.delete(r); db.session.commit(); audit('Financeiro',id,'EXCLUÍDO',''); flash('Lançamento excluído.','ok'); return redirect(url_for('financeiro'))

@app.route('/financeiro/exportar')
@login_required
@allow('Administrador')
def exportar_financeiro():
    wb=Workbook(); ws=wb.active; ws.title='Resumo'; rows=Financeiro.query.order_by(Financeiro.data,Financeiro.id).all()
    ent=sum(x.valor or 0 for x in rows if x.tipo=='Entrada' and x.status=='Pago/Recebido'); sai=sum(x.valor or 0 for x in rows if x.tipo=='Saída' and x.status=='Pago/Recebido')
    ws.append(['Gerenciamento BR Soluções - Financeiro 13.2']); ws.append(['Entradas',ent]); ws.append(['Saídas',sai]); ws.append(['Saldo',ent-sai])
    ws['A1'].font=Font(bold=True,size=16,color='FFFFFF'); ws['A1'].fill=PatternFill('solid',fgColor='24324A'); ws.merge_cells('A1:D1')
    ws.column_dimensions['A'].width=24; ws.column_dimensions['B'].width=18
    ws.append([]); ws.append(['Mês','Entradas','Saídas','Saldo'])
    mensal={}
    for r in rows:
        mes=(r.data or '')[:7] or 'Sem data'; mensal.setdefault(mes,[0.0,0.0])
        if r.status=='Pago/Recebido': mensal[mes][0 if r.tipo=='Entrada' else 1]+=r.valor or 0
    for mes in sorted(mensal): ws.append([mes,mensal[mes][0],mensal[mes][1],mensal[mes][0]-mensal[mes][1]])
    if mensal:
        min_row=7; max_row=6+len(mensal)
        chart=BarChart(); chart.title='Entradas e saídas por mês'; chart.y_axis.title='Valor (R$)'; chart.x_axis.title='Mês'; chart.height=8; chart.width=16
        chart.add_data(Reference(ws,min_col=2,max_col=3,min_row=6,max_row=max_row),titles_from_data=True); chart.set_categories(Reference(ws,min_col=1,min_row=min_row,max_row=max_row)); ws.add_chart(chart,'F2')
        line=LineChart(); line.title='Evolução do saldo mensal'; line.y_axis.title='Saldo (R$)'; line.height=8; line.width=16
        line.add_data(Reference(ws,min_col=4,min_row=6,max_row=max_row),titles_from_data=True); line.set_categories(Reference(ws,min_col=1,min_row=min_row,max_row=max_row)); ws.add_chart(line,'F18')
    for nome,tipo in [('Entradas','Entrada'),('Saídas','Saída')]:
        sh=wb.create_sheet(nome); sh.append(['Data','Tipo','Descrição','Origem','Construtora','Empreendimento','Categoria','Valor','Status','Forma de pagamento','Vencimento','Documento','Observações'])
        for r in rows:
            if r.tipo==tipo: sh.append([r.data,r.tipo,r.descricao,r.origem,r.construtora,r.empreendimento,r.categoria,r.valor,r.status,r.forma_pagamento,r.vencimento,r.documento,r.observacoes])
        sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
        for c in sh[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='0B3B5A')
    out=io.BytesIO(); wb.save(out); out.seek(0); return send_file(out,as_attachment=True,download_name='Financeiro_Gerenciamento_BR_Solucoes_13_2.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/empreendimentos',methods=['GET','POST'])
@login_required
@allow('Administrador','Líder')
def empreendimentos():
    if request.method=='POST':
        if not Empreendimento.query.filter_by(nome=request.form.get('nome')).first(): db.session.add(Empreendimento(nome=request.form.get('nome'),construtora=request.form.get('construtora') or 'Prestes',data_entrega=request.form.get('data_entrega'),observacoes=request.form.get('observacoes')));db.session.commit()
        return redirect(url_for('empreendimentos'))
    return render_template('empreendimentos.html',rows=Empreendimento.query.order_by(Empreendimento.nome).all(),construtoras=Construtora.query.order_by(Construtora.nome).all())

@app.route('/usuarios',methods=['GET','POST'])
@login_required
@allow('Administrador')
def usuarios():
    if request.method=='POST':
        email=request.form.get('email','').lower().strip()
        if Usuario.query.filter_by(email=email).first(): flash('E-mail já cadastrado.','erro')
        else: db.session.add(Usuario(nome=request.form.get('nome'),email=email,senha=ph(request.form.get('senha')),perfil=request.form.get('perfil'),construtora=request.form.get('construtora'),ativo=True));db.session.commit();flash('Usuário cadastrado.','ok')
        return redirect(url_for('usuarios'))
    return render_template('usuarios.html',rows=Usuario.query.order_by(Usuario.nome).all(),construtoras=Construtora.query.order_by(Construtora.nome).all())
@app.post('/usuarios/<int:id>/alternar')
@login_required
@allow('Administrador')
def alternar_usuario(id):
    u=Usuario.query.get_or_404(id)
    if u.id==session.get('user_id'): flash('Você não pode desativar seu próprio usuário.','erro')
    else: u.ativo=not u.ativo;db.session.commit()
    return redirect(url_for('usuarios'))

@app.route('/exportar')
@login_required
def exportar():
    empreendimento_filtro=(request.args.get('empreendimento') or '').strip()
    qry=Chamado.query
    if empreendimento_filtro:
        qry=qry.filter(Chamado.empreendimento==empreendimento_filtro)
    rows=qry.order_by(Chamado.empreendimento,Chamado.status,Chamado.data_recebido,Chamado.id).all()

    wb=Workbook()
    resumo=wb.active
    resumo.title='Resumo Geral'

    azul='0B3B5A'; azul_claro='D9EAF4'; branco='FFFFFF'; cinza='E5E7EB'; laranja='F06B22'
    borda=Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    status_cores={
      'aberto':'BDD7EE','aberta':'BDD7EE','em execucao':'FFF2CC','em execução':'FFF2CC','em andamento':'FFF2CC',
      'aguardando material':'FCE4D6','aguardando morador':'E4DFEC','aguardando cliente':'E4DFEC',
      'finalizado':'C6E0B4','finalizada':'C6E0B4','concluido':'C6E0B4','concluído':'C6E0B4',
      'improcedente':'F4CCCC','procedente':'D9EAD3','em analise':'D9D9D9','em análise':'D9D9D9','cancelado':'F4CCCC','cancelada':'F4CCCC'
    }
    def safe_sheet_name(name,used):
        base=re.sub(r'[\/*?:\[\]]','-',(name or 'Sem empreendimento')).strip()[:31] or 'Sem empreendimento'
        candidate=base; n=2
        while candidate in used:
            suffix=f' ({n})'; candidate=(base[:31-len(suffix)]+suffix); n+=1
        used.add(candidate); return candidate
    def style_header(ws,row=1):
        for cell in ws[row]:
            cell.fill=PatternFill('solid',fgColor=azul); cell.font=Font(color=branco,bold=True); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=borda
        ws.row_dimensions[row].height=28
    def format_status(cell):
        key=norm_header(cell.value)
        color=status_cores.get(key)
        if color: cell.fill=PatternFill('solid',fgColor=color)
        cell.font=Font(bold=True); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=borda

    # Resumo geral
    resumo.merge_cells('A1:F1'); resumo['A1']='BR SmartSAT — Resumo de Chamados'
    resumo['A1'].fill=PatternFill('solid',fgColor=azul); resumo['A1'].font=Font(color=branco,bold=True,size=16); resumo['A1'].alignment=Alignment(horizontal='center'); resumo.row_dimensions[1].height=30
    resumo.append([])
    resumo.append(['Indicador','Quantidade'])
    total=len(rows)
    resumo.append(['Total de SATs',total])
    status_counts={}
    emp_counts={}
    for r in rows:
        st=(r.status or 'Sem status').strip(); status_counts[st]=status_counts.get(st,0)+1
        emp=(r.empreendimento or 'Sem empreendimento').strip(); emp_counts[emp]=emp_counts.get(emp,0)+1
    for st,count in sorted(status_counts.items(),key=lambda x:x[0].lower()): resumo.append([st,count])
    style_header(resumo,3)
    for row in resumo.iter_rows(min_row=4,max_row=3+1+len(status_counts),min_col=1,max_col=2):
        for c in row: c.border=borda
        if row[0].value!='Total de SATs': format_status(row[0])
    start=6+len(status_counts)
    resumo.cell(start,1,'Empreendimento'); resumo.cell(start,2,'Quantidade')
    for c in resumo[start]: c.fill=PatternFill('solid',fgColor=azul); c.font=Font(color=branco,bold=True); c.border=borda
    for i,(emp,count) in enumerate(sorted(emp_counts.items(),key=lambda x:x[0].lower()),start=start+1):
        resumo.cell(i,1,emp); resumo.cell(i,2,count); resumo.cell(i,1).border=borda; resumo.cell(i,2).border=borda
    resumo.column_dimensions['A'].width=34; resumo.column_dimensions['B'].width=16; resumo.freeze_panes='A4'

    headers=['SAT','Data de abertura','Empreendimento','Unidade','Morador / Solicitante','Contato','Problema / Solicitação','Categoria','Classificação','Status','Responsável','Observações','Última atualização']
    grouped={}
    for r in rows: grouped.setdefault((r.empreendimento or 'Sem empreendimento').strip(),[]).append(r)
    used={'Resumo Geral'}
    for emp,items in sorted(grouped.items(),key=lambda x:x[0].lower()):
        ws=wb.create_sheet(safe_sheet_name(emp,used))
        ws.merge_cells('A1:M1'); ws['A1']=f'AGENDA SATS - {emp.upper()}'; ws['A1'].fill=PatternFill('solid',fgColor=laranja); ws['A1'].font=Font(color=branco,bold=True,size=15); ws['A1'].alignment=Alignment(horizontal='center'); ws.row_dimensions[1].height=28
        ws.append(headers); style_header(ws,2)
        for cell in ws[2]: cell.fill=PatternFill('solid',fgColor=laranja)
        for r in items:
            ws.append([
              r.sat or r.id, excel_date(r.data_recebido), r.empreendimento or '', r.unidade or '', r.solicitante or '', r.contato or '',
              r.problema or '', r.categoria or '', r.classificacao or '', r.status or 'Sem status', r.responsavel or '', r.observacoes or '',
              r.atualizado_em.strftime('%d/%m/%Y %H:%M') if r.atualizado_em else ''
            ])
            rr=ws.max_row
            for c in ws[rr]: c.border=borda; c.alignment=Alignment(vertical='top',wrap_text=True)
            format_status(ws.cell(rr,10))
        ws.auto_filter.ref=f'A2:M{ws.max_row}'
        ws.freeze_panes='A3'
        widths=[13,16,28,16,26,18,48,34,18,20,24,40,20]
        for idx,width in enumerate(widths,1): ws.column_dimensions[get_column_letter(idx)].width=width
        for rr in range(3,ws.max_row+1): ws.row_dimensions[rr].height=42
        ws.sheet_view.showGridLines=False

    out=io.BytesIO(); wb.save(out); out.seek(0)
    nome='Chamados_BR_SmartSAT_por_empreendimento.xlsx' if not empreendimento_filtro else f'Chamados_{secure_filename(empreendimento_filtro)}.xlsx'
    return send_file(out,as_attachment=True,download_name=nome,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/agenda',methods=['GET','POST'])
@login_required
@allow('Administrador','Líder','Equipe','Técnico','Construtora')
def agenda():
    if request.method=='POST':
        if session.get('perfil') not in ('Administrador','Líder'): abort(403)
        chamado_id=request.form.get('chamado_id',type=int)
        if not chamado_id: abort(400)
        c=Chamado.query.get_or_404(chamado_id)
        fluxo=(request.form.get('fluxo') or 'agendar_direto').strip()
        lider_id=request.form.get('lider_id',type=int)
        lider_usuario=Usuario.query.get(lider_id) if lider_id else None
        if lider_usuario and lider_usuario.perfil not in ('Líder','Administrador'): abort(400)
        if session.get('perfil')=='Líder': lider_usuario=Usuario.query.get(session.get('user_id'))
        equipe_id=request.form.get('equipe_id',type=int)
        equipe_usuario=Usuario.query.get(equipe_id) if equipe_id else None
        if equipe_usuario and equipe_usuario.perfil not in ('Equipe','Técnico','Líder'): abort(400)
        if fluxo=='encaminhar_lider':
            if session.get('perfil')!='Administrador' or not lider_usuario: abort(400)
            status='Aguardando Líder'; equipe_usuario=None; lider_confirmou=False; agendado_por=None
        else:
            if not equipe_usuario: abort(400)
            status='Agendada pelo Administrador' if session.get('perfil')=='Administrador' else 'Agendada pelo Líder'
            lider_confirmou=True; agendado_por=session.get('nome')
        a=Agenda(chamado_id=c.id,construtora=c.construtora or 'Prestes',empreendimento=c.empreendimento,
            lider=lider_usuario.nome if lider_usuario else session.get('nome'),lider_id=lider_usuario.id if lider_usuario else session.get('user_id'),lider_email=lider_usuario.email if lider_usuario else session.get('email'),
            equipe=equipe_usuario.nome if equipe_usuario else None,equipe_id=equipe_usuario.id if equipe_usuario else None,equipe_email=equipe_usuario.email if equipe_usuario else None,
            prazo_contato=request.form.get('prazo_contato'),data_agendada=request.form.get('data_agendada'),hora=request.form.get('hora'),periodo=request.form.get('periodo'),status=status,tipo_fluxo=fluxo,
            contato_status=request.form.get('contato_status'),observacao_interna=request.form.get('observacao_interna'),observacao_construtora=request.form.get('observacao_construtora'),criado_por=session.get('nome'),agendado_por=agendado_por,
            lider_confirmou=lider_confirmou,lider_visualizou=(session.get('perfil')=='Líder'),confirmado_em=datetime.utcnow() if lider_confirmou else None)
        db.session.add(a); db.session.commit()
        if fluxo=='encaminhar_lider':
            audit('Agenda',a.id,'ENVIADO_AO_LIDER',f'SAT {c.sat or c.id} enviada para {a.lider}')
            flash('SAT enviada ao líder. Ela já aparece como notificação pendente.','ok')
        else:
            audit('Agenda',a.id,'AGENDADO_DIRETAMENTE',f'SAT {c.sat or c.id} | Equipe: {a.equipe} | {a.data_agendada} {a.hora or a.periodo or ""}')
            flash('Agendamento criado e enviado diretamente para a equipe.','ok')
        return redirect(url_for('agenda'))
    qry=Agenda.query
    if session.get('perfil')=='Construtora':
        qry=qry.filter(Agenda.construtora==session.get('construtora'))
    elif session.get('perfil')=='Líder':
        qry=qry.filter(agenda_destino_filter('lider'))
    elif session.get('perfil') in ('Equipe','Técnico'):
        qry=qry.filter(agenda_destino_filter('equipe'))
    mes=request.args.get('mes'); emp=request.args.get('empreendimento'); status=request.args.get('status')
    if mes: qry=qry.filter(Agenda.data_agendada.like(mes+'%'))
    if emp: qry=qry.filter_by(empreendimento=emp)
    if status: qry=qry.filter_by(status=status)
    rows=qry.order_by(Agenda.data_agendada.asc(),Agenda.hora.asc(),Agenda.id.desc()).all()
    ids=[x.chamado_id for x in rows]; chamados_map={c.id:c for c in Chamado.query.filter(Chamado.id.in_(ids)).all()} if ids else {}
    try:
        cal_year,cal_month=(map(int,mes.split('-')) if mes else (date.today().year,date.today().month))
    except Exception: cal_year,cal_month=date.today().year,date.today().month
    cal=calendar.Calendar(firstweekday=6)
    weeks=cal.monthdayscalendar(cal_year,cal_month)
    by_day={}
    for a in rows:
        if a.data_agendada and a.data_agendada.startswith(f'{cal_year:04d}-{cal_month:02d}-'):
            try: by_day.setdefault(int(a.data_agendada[-2:]),[]).append(a)
            except: pass
    prev_month=cal_month-1 or 12; prev_year=cal_year-1 if cal_month==1 else cal_year
    next_month=cal_month+1 if cal_month<12 else 1; next_year=cal_year+1 if cal_month==12 else cal_year
    avisos=[]
    if session.get('perfil')=='Líder':
        avisos=Agenda.query.filter(agenda_destino_filter('lider'),Agenda.lider_confirmou==False,Agenda.status.in_(['Aguardando Líder','Aguardando agendamento do líder'])).order_by(Agenda.criado_em.desc()).all()
    return render_template('agenda.html',rows=rows,chamados_map=chamados_map,chamados=Chamado.query.order_by(Chamado.id.desc()).all(),lideres=Usuario.query.filter(Usuario.perfil.in_(['Líder','Administrador'])).all(),equipes=Usuario.query.filter(Usuario.perfil.in_(['Equipe','Técnico','Líder'])).all(),emps=Empreendimento.query.order_by(Empreendimento.nome).all(),weeks=weeks,by_day=by_day,avisos=avisos,cal_year=cal_year,cal_month=cal_month,month_name=['','Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'][cal_month],prev_mes=f'{prev_year:04d}-{prev_month:02d}',next_mes=f'{next_year:04d}-{next_month:02d}')



@app.post('/agenda/<int:id>/confirmar-lider')
@login_required
@allow('Líder','Administrador')
def agenda_confirmar_lider(id):
    a=Agenda.query.get_or_404(id)
    if session.get('perfil')=='Líder' and not (a.lider_id==session.get('user_id') or (a.lider or '').lower()==session.get('nome','').lower()): abort(403)
    if session.get('perfil')=='Líder':
        a.lider_visualizou=True
        db.session.commit()
        flash('Confira a sugestão, confirme a data com o cliente e escolha a equipe.','ok')
        return redirect(url_for('agenda_editar',id=id))
    a.lider_visualizou=True; a.lider_confirmou=True
    if a.status in ('Aguardando contato','Aguardando agendamento do líder','Aguardando confirmação'): a.status='Agendada'
    if not a.contato_status: a.contato_status='Cliente confirmou'
    db.session.commit(); audit('Agenda',a.id,'CONFIRMADO_PELO_ADMIN',f'Data {a.data_agendada or "a definir"} {a.hora or a.periodo or ""}')
    flash('Agendamento confirmado.','ok')
    return redirect(url_for('agenda'))

@app.get('/agenda/<int:id>/abrir-aviso')
@login_required
@allow('Líder','Administrador')
def agenda_abrir_aviso(id):
    a=Agenda.query.get_or_404(id)
    if session.get('perfil')=='Líder' and not (a.lider_id==session.get('user_id') or (a.lider or '').lower()==session.get('nome','').lower()): abort(403)
    a.lider_visualizou=True; db.session.commit()
    return redirect(url_for('agenda_editar',id=id))

@app.get('/api/agenda/versao')
@login_required
def api_agenda_versao():
    qry=Agenda.query
    if session.get('perfil')=='Líder':
        qry=qry.filter(agenda_destino_filter('lider'))
    elif session.get('perfil') in ('Equipe','Técnico'):
        qry=qry.filter(agenda_destino_filter('equipe'))
    elif session.get('perfil')=='Construtora':
        qry=qry.filter(Agenda.construtora==session.get('construtora'))
    last=qry.order_by(Agenda.atualizado_em.desc(),Agenda.id.desc()).first()
    return jsonify({'total':qry.count(),'ultimo':last.atualizado_em.isoformat() if last and last.atualizado_em else '', 'id':last.id if last else 0})

@app.route('/agenda/exportar')
@login_required
@allow('Administrador','Líder','Construtora')
def exportar_agenda():
    qry=Agenda.query
    if session.get('perfil')=='Construtora':
        qry=qry.filter(Agenda.construtora==session.get('construtora'))
    mes=request.args.get('mes','').strip(); emp=request.args.get('empreendimento','').strip(); status=request.args.get('status','').strip()
    if mes: qry=qry.filter(Agenda.data_agendada.like(mes+'%'))
    if emp: qry=qry.filter_by(empreendimento=emp)
    if status: qry=qry.filter_by(status=status)
    agendas=qry.order_by(Agenda.empreendimento,Agenda.data_agendada,Agenda.hora,Agenda.id).all()
    chamado_ids=[a.chamado_id for a in agendas]
    cmap={c.id:c for c in Chamado.query.filter(Chamado.id.in_(chamado_ids)).all()} if chamado_ids else {}

    wb=Workbook(); resumo=wb.active; resumo.title='Resumo Geral'
    azul='172033'; azul2='24324A'; cobre='D4884A'; branco='FFFFFF'; claro='F4F1EB'; verde='24856B'; vermelho='B94C55'; palette=['F06B22','2E75B6','70AD47','7030A0','C55A11','4472C4','A5A5A5']
    borda=Border(left=Side(style='thin',color='DDE2E8'),right=Side(style='thin',color='DDE2E8'),top=Side(style='thin',color='DDE2E8'),bottom=Side(style='thin',color='DDE2E8'))
    resumo.merge_cells('A1:H1'); resumo['A1']='Gerenciamento BR Soluções — Agenda de Atendimentos'
    resumo['A1'].fill=PatternFill('solid',fgColor=azul); resumo['A1'].font=Font(color=branco,bold=True,size=16); resumo['A1'].alignment=Alignment(horizontal='center'); resumo.row_dimensions[1].height=30
    resumo.append([]); resumo.append(['Indicador','Quantidade'])
    resumo.append(['Total de agendamentos',len(agendas)])
    resumo.append(['Agendados',sum(1 for a in agendas if a.status=='Agendada')])
    resumo.append(['Reagendados',sum(1 for a in agendas if a.status=='Reagendada')])
    resumo.append(['Realizados',sum(1 for a in agendas if a.status=='Atendimento realizado')])
    resumo.append(['Aguardando contato',sum(1 for a in agendas if a.status=='Aguardando contato')])
    for cell in resumo[3]: cell.fill=PatternFill('solid',fgColor=azul2); cell.font=Font(color=branco,bold=True); cell.border=borda
    for row in resumo.iter_rows(min_row=4,max_row=8,min_col=1,max_col=2):
        for cell in row: cell.border=borda
    resumo.column_dimensions['A'].width=28; resumo.column_dimensions['B'].width=16; resumo.sheet_view.showGridLines=False

    headers=['Data','Hora','Período','SAT','Torre / Apartamento','Descrição','Cliente','Contato','Construtora','Empreendimento','Líder','Equipe / Técnico','Status','Observação para construtora']
    grouped={}
    for a in agendas: grouped.setdefault(a.empreendimento or 'Sem empreendimento',[]).append(a)
    used={'Resumo Geral'}
    for color_index,(nome,items) in enumerate(sorted(grouped.items(),key=lambda x:x[0].lower())):
        base=re.sub(r'[\/*?:\[\]]',' ',nome).strip()[:31] or 'Agenda'
        sheet=base; i=2
        while sheet in used:
            suffix=f' {i}'; sheet=(base[:31-len(suffix)]+suffix); i+=1
        used.add(sheet); ws=wb.create_sheet(sheet); emp_color=palette[color_index % len(palette)]
        ws.merge_cells('A1:N1'); ws['A1']=f'AGENDA SATS - {nome.upper()}'; ws['A1'].fill=PatternFill('solid',fgColor=emp_color); ws['A1'].font=Font(color=branco,bold=True,size=15); ws['A1'].alignment=Alignment(horizontal='center'); ws.row_dimensions[1].height=28
        ws.append(headers)
        for cell in ws[2]: cell.fill=PatternFill('solid',fgColor=emp_color); cell.font=Font(color=branco,bold=True); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=borda
        for a in items:
            c=cmap.get(a.chamado_id)
            ws.append([
                excel_date(a.data_agendada),a.hora or '',a.periodo or '',(c.sat if c else a.chamado_id),
                (c.unidade if c else ''),(c.problema if c else ''),(c.solicitante if c else ''),(c.contato if c else ''),
                a.construtora or '',a.empreendimento or '',a.lider or '',a.equipe or '',a.status or '',a.observacao_construtora or ''
            ])
            rr=ws.max_row
            fill='FFFFFF' if rr%2 else claro
            for cell in ws[rr]: cell.fill=PatternFill('solid',fgColor=fill); cell.border=borda; cell.alignment=Alignment(vertical='top',wrap_text=True)
            st=ws.cell(rr,13)
            if a.status=='Atendimento realizado': st.fill=PatternFill('solid',fgColor='DDF2EA'); st.font=Font(color=verde,bold=True)
            elif a.status in ('Cancelada','Cliente não respondeu'): st.fill=PatternFill('solid',fgColor='F7E3E5'); st.font=Font(color=vermelho,bold=True)
            elif a.status in ('Agendada','Reagendada'): st.fill=PatternFill('solid',fgColor='F8E9D9'); st.font=Font(color='9B5A24',bold=True)
        ws.auto_filter.ref=f'A2:N{max(ws.max_row,2)}'; ws.freeze_panes='A3'; ws.sheet_view.showGridLines=False
        widths=[13,10,14,13,20,48,27,18,24,28,24,24,22,42]
        for idx,width in enumerate(widths,1): ws.column_dimensions[get_column_letter(idx)].width=width
        for rr in range(3,ws.max_row+1): ws.row_dimensions[rr].height=44
    out=io.BytesIO(); wb.save(out); out.seek(0)
    filename='Agenda_Gerenciamento_BR_Solucoes.xlsx' if not emp else f'Agenda_{secure_filename(emp)}.xlsx'
    return send_file(out,as_attachment=True,download_name=filename,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/agenda/<int:id>/editar',methods=['GET','POST'])
@login_required
@allow('Administrador','Líder')
def agenda_editar(id):
    a=Agenda.query.get_or_404(id)
    if session.get('perfil')=='Líder' and not (a.lider_id==session.get('user_id') or (a.lider_email and a.lider_email.lower()==(session.get('email') or '').lower()) or (a.lider or '').lower()==session.get('nome','').lower()): abort(403)
    if request.method=='POST':
        if session.get('perfil')=='Líder':
            equipe_id=request.form.get('equipe_id',type=int)
            equipe_usuario=Usuario.query.get(equipe_id) if equipe_id else None
            if not equipe_usuario or equipe_usuario.perfil not in ('Equipe','Técnico','Líder'): abort(400)
            a.equipe=equipe_usuario.nome; a.equipe_id=equipe_usuario.id; a.equipe_email=equipe_usuario.email
            a.data_agendada=request.form.get('data_agendada'); a.hora=request.form.get('hora'); a.periodo=request.form.get('periodo')
            a.contato_status=request.form.get('contato_status') or 'Cliente confirmou'; a.observacao_interna=request.form.get('observacao_interna')
            a.lider_visualizou=True; a.lider_confirmou=True; a.confirmado_em=datetime.utcnow(); a.agendado_por=session.get('nome'); a.tipo_fluxo='encaminhar_lider'; a.status='Agendada pelo Líder'
            db.session.commit(); audit('Agenda',id,'AGENDADO_PELO_LIDER',f'{a.data_agendada} {a.hora or a.periodo or ""} | Equipe: {a.equipe}')
            flash('Agenda confirmada e enviada automaticamente para a equipe.','ok')
        else:
            for f in ['prazo_contato','data_agendada','hora','periodo','status','contato_status','observacao_interna','observacao_construtora']: setattr(a,f,request.form.get(f))
            lider_id=request.form.get('lider_id',type=int)
            if lider_id:
                u=Usuario.query.get(lider_id)
                if u: a.lider_id=u.id; a.lider=u.nome; a.lider_email=u.email
            equipe_id=request.form.get('equipe_id',type=int)
            if equipe_id:
                u=Usuario.query.get(equipe_id)
                if u: a.equipe_id=u.id; a.equipe=u.nome; a.equipe_email=u.email
            db.session.commit(); audit('Agenda',id,'EDITADO',a.status); flash('Agenda atualizada.','ok')
        return redirect(url_for('agenda'))
    return render_template('agenda_editar.html',a=a,lideres=Usuario.query.filter(Usuario.perfil.in_(['Líder','Administrador'])).all(),equipes=Usuario.query.filter(Usuario.perfil.in_(['Equipe','Técnico','Líder'])).all())

@app.post('/agenda/<int:id>/excluir')
@login_required
@allow('Administrador')
def agenda_excluir(id):
    a=Agenda.query.get_or_404(id); db.session.delete(a); db.session.commit(); audit('Agenda',id,'EXCLUÍDO',''); flash('Agendamento excluído.','ok'); return redirect(url_for('agenda'))

@app.route('/construtoras',methods=['GET','POST'])
@login_required
@allow('Administrador')
def construtoras():
    if request.method=='POST':
        nome=request.form.get('nome','').strip()
        if nome and not Construtora.query.filter(func.lower(Construtora.nome)==nome.lower()).first(): db.session.add(Construtora(nome=nome)); db.session.commit(); flash('Construtora cadastrada.','ok')
        return redirect(url_for('construtoras'))
    return render_template('construtoras.html',rows=Construtora.query.order_by(Construtora.nome).all())

@app.route('/api/analisar',methods=['POST'])
@login_required
def api_analisar():
    d=request.get_json(force=True);cat=d.get('categoria') or guess_category(d.get('problema'));return jsonify({'categoria':cat,**warranty_analysis(d.get('empreendimento'),cat,d.get('data_entrega'),d.get('data_abertura'))})

with app.app_context(): init_db()
if __name__=='__main__': app.run(host='0.0.0.0',port=int(os.environ.get('PORT',5000)),debug=False)
